from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from typing import List, Optional
from jose import jwt, JWTError
from passlib.context import CryptContext

from database import SessionLocal, engine, Base
from models import User, Income, Expense, Budget, SavingsGoal
from schemas import (
    UserCreate, UserResponse, Token,
    IncomeCreate, IncomeUpdate, IncomeResponse,
    ExpenseCreate, ExpenseUpdate, ExpenseResponse,
    BudgetCreate, BudgetUpdate, BudgetResponse,
    SavingsGoalCreate, SavingsGoalUpdate, SavingsGoalResponse,
    DashboardStats, CategoryStats
)

# Créer les tables
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Finance Management API", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration sécurité
SECRET_KEY = "your-secret-key-change-in-production"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")


# Dépendances
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    return pwd_context.hash(password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: int = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise credentials_exception
    return user


# Routes d'authentification
@app.post("/register", response_model=UserResponse)
def register(user: UserCreate, db: Session = Depends(get_db)):
    try:
        # Vérifier si l'utilisateur existe déjà
        db_user = db.query(User).filter(User.email == user.email).first()
        if db_user:
            raise HTTPException(status_code=400, detail="Email already registered")
        
        # Valider le mot de passe
        if len(user.password) < 6:
            raise HTTPException(status_code=400, detail="Le mot de passe doit contenir au moins 6 caractères")
        
        hashed_password = get_password_hash(user.password)
        db_user = User(
            email=user.email,
            hashed_password=hashed_password,
            full_name=user.full_name
        )
        db.add(db_user)
        db.commit()
        db.refresh(db_user)
        return db_user
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'inscription: {str(e)}")


@app.post("/token", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.id}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/users/me", response_model=UserResponse)
def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user


# Routes pour les revenus
@app.post("/incomes", response_model=IncomeResponse)
def create_income(income: IncomeCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_income = Income(**income.dict(), user_id=current_user.id)
    db.add(db_income)
    db.commit()
    db.refresh(db_income)
    return db_income


@app.get("/incomes", response_model=List[IncomeResponse])
def read_incomes(skip: int = 0, limit: int = 100, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    incomes = db.query(Income).filter(Income.user_id == current_user.id).offset(skip).limit(limit).all()
    return incomes


@app.get("/incomes/{income_id}", response_model=IncomeResponse)
def read_income(income_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    income = db.query(Income).filter(Income.id == income_id, Income.user_id == current_user.id).first()
    if income is None:
        raise HTTPException(status_code=404, detail="Income not found")
    return income


@app.put("/incomes/{income_id}", response_model=IncomeResponse)
def update_income(income_id: int, income: IncomeUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_income = db.query(Income).filter(Income.id == income_id, Income.user_id == current_user.id).first()
    if db_income is None:
        raise HTTPException(status_code=404, detail="Income not found")
    for key, value in income.dict(exclude_unset=True).items():
        setattr(db_income, key, value)
    db.commit()
    db.refresh(db_income)
    return db_income


@app.delete("/incomes/{income_id}")
def delete_income(income_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    income = db.query(Income).filter(Income.id == income_id, Income.user_id == current_user.id).first()
    if income is None:
        raise HTTPException(status_code=404, detail="Income not found")
    db.delete(income)
    db.commit()
    return {"message": "Income deleted successfully"}


# Routes pour les dépenses
@app.post("/expenses", response_model=ExpenseResponse)
def create_expense(expense: ExpenseCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_expense = Expense(**expense.dict(), user_id=current_user.id)
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    return db_expense


@app.get("/expenses", response_model=List[ExpenseResponse])
def read_expenses(skip: int = 0, limit: int = 100, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    expenses = db.query(Expense).filter(Expense.user_id == current_user.id).offset(skip).limit(limit).all()
    return expenses


@app.get("/expenses/{expense_id}", response_model=ExpenseResponse)
def read_expense(expense_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id, Expense.user_id == current_user.id).first()
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    return expense


@app.put("/expenses/{expense_id}", response_model=ExpenseResponse)
def update_expense(expense_id: int, expense: ExpenseUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_expense = db.query(Expense).filter(Expense.id == expense_id, Expense.user_id == current_user.id).first()
    if db_expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    for key, value in expense.dict(exclude_unset=True).items():
        setattr(db_expense, key, value)
    db.commit()
    db.refresh(db_expense)
    return db_expense


@app.delete("/expenses/{expense_id}")
def delete_expense(expense_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id, Expense.user_id == current_user.id).first()
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    db.delete(expense)
    db.commit()
    return {"message": "Expense deleted successfully"}


# Routes pour les budgets
@app.post("/budgets", response_model=BudgetResponse)
def create_budget(budget: BudgetCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_budget = Budget(**budget.dict(), user_id=current_user.id)
    db.add(db_budget)
    db.commit()
    db.refresh(db_budget)
    return db_budget


@app.get("/budgets", response_model=List[BudgetResponse])
def read_budgets(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).all()
    return budgets


@app.put("/budgets/{budget_id}", response_model=BudgetResponse)
def update_budget(budget_id: int, budget: BudgetUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == current_user.id).first()
    if db_budget is None:
        raise HTTPException(status_code=404, detail="Budget not found")
    for key, value in budget.dict(exclude_unset=True).items():
        setattr(db_budget, key, value)
    db.commit()
    db.refresh(db_budget)
    return db_budget


@app.delete("/budgets/{budget_id}")
def delete_budget(budget_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == current_user.id).first()
    if budget is None:
        raise HTTPException(status_code=404, detail="Budget not found")
    db.delete(budget)
    db.commit()
    return {"message": "Budget deleted successfully"}


# Routes pour les objectifs d'épargne
@app.post("/savings-goals", response_model=SavingsGoalResponse)
def create_savings_goal(goal: SavingsGoalCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_goal = SavingsGoal(**goal.dict(), user_id=current_user.id)
    db.add(db_goal)
    db.commit()
    db.refresh(db_goal)
    return db_goal


@app.get("/savings-goals", response_model=List[SavingsGoalResponse])
def read_savings_goals(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    goals = db.query(SavingsGoal).filter(SavingsGoal.user_id == current_user.id).all()
    return goals


@app.put("/savings-goals/{goal_id}", response_model=SavingsGoalResponse)
def update_savings_goal(goal_id: int, goal: SavingsGoalUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_goal = db.query(SavingsGoal).filter(SavingsGoal.id == goal_id, SavingsGoal.user_id == current_user.id).first()
    if db_goal is None:
        raise HTTPException(status_code=404, detail="Savings goal not found")
    for key, value in goal.dict(exclude_unset=True).items():
        setattr(db_goal, key, value)
    db.commit()
    db.refresh(db_goal)
    return db_goal


@app.delete("/savings-goals/{goal_id}")
def delete_savings_goal(goal_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    goal = db.query(SavingsGoal).filter(SavingsGoal.id == goal_id, SavingsGoal.user_id == current_user.id).first()
    if goal is None:
        raise HTTPException(status_code=404, detail="Savings goal not found")
    db.delete(goal)
    db.commit()
    return {"message": "Savings goal deleted successfully"}


# Route pour le dashboard
@app.get("/dashboard", response_model=DashboardStats)
def get_dashboard(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    from sqlalchemy import func, extract
    from datetime import datetime
    
    # Revenus totaux
    total_income = db.query(func.sum(Income.amount)).filter(Income.user_id == current_user.id).scalar() or 0
    
    # Dépenses totales
    total_expenses = db.query(func.sum(Expense.amount)).filter(Expense.user_id == current_user.id).scalar() or 0
    
    # Solde actuel
    balance = total_income - total_expenses
    
    # Revenus du mois en cours
    current_month = datetime.now().month
    current_year = datetime.now().year
    monthly_income = db.query(func.sum(Income.amount)).filter(
        Income.user_id == current_user.id,
        extract('month', Income.date) == current_month,
        extract('year', Income.date) == current_year
    ).scalar() or 0
    
    # Dépenses du mois en cours
    monthly_expenses = db.query(func.sum(Expense.amount)).filter(
        Expense.user_id == current_user.id,
        extract('month', Expense.date) == current_month,
        extract('year', Expense.date) == current_year
    ).scalar() or 0
    
    # Dépenses par catégorie
    category_expenses = db.query(
        Expense.category,
        func.sum(Expense.amount).label('total')
    ).filter(
        Expense.user_id == current_user.id,
        extract('month', Expense.date) == current_month,
        extract('year', Expense.date) == current_year
    ).group_by(Expense.category).all()
    
    category_stats = [CategoryStats(category=cat, amount=float(amt)) for cat, amt in category_expenses]
    
    # Évolution mensuelle (6 derniers mois)
    monthly_evolution = []
    for i in range(6):
        month = current_month - i
        year = current_year
        if month <= 0:
            month += 12
            year -= 1
        
        month_income = db.query(func.sum(Income.amount)).filter(
            Income.user_id == current_user.id,
            extract('month', Income.date) == month,
            extract('year', Income.date) == year
        ).scalar() or 0
        
        month_expenses = db.query(func.sum(Expense.amount)).filter(
            Expense.user_id == current_user.id,
            extract('month', Expense.date) == month,
            extract('year', Expense.date) == year
        ).scalar() or 0
        
        monthly_evolution.append({
            "month": f"{year}-{month:02d}",
            "income": float(month_income),
            "expenses": float(month_expenses)
        })
    
    monthly_evolution.reverse()
    
    # Indicateur de santé financière
    if monthly_expenses > 0:
        savings_rate = ((monthly_income - monthly_expenses) / monthly_income * 100) if monthly_income > 0 else 0
        if savings_rate >= 20:
            health_status = "excellent"
        elif savings_rate >= 10:
            health_status = "bon"
        elif savings_rate >= 0:
            health_status = "moyen"
        else:
            health_status = "critique"
    else:
        health_status = "excellent"
    
    return DashboardStats(
        balance=float(balance),
        total_income=float(total_income),
        total_expenses=float(total_expenses),
        monthly_income=float(monthly_income),
        monthly_expenses=float(monthly_expenses),
        category_stats=category_stats,
        monthly_evolution=monthly_evolution,
        health_status=health_status
    )


# Route pour export CSV
@app.get("/export/csv")
def export_csv(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # En-têtes
    writer.writerow(["Type", "Date", "Catégorie", "Montant", "Commentaire"])
    
    # Revenus
    incomes = db.query(Income).filter(Income.user_id == current_user.id).all()
    for income in incomes:
        writer.writerow(["Revenu", income.date.strftime("%Y-%m-%d"), income.category, income.amount, ""])
    
    # Dépenses
    expenses = db.query(Expense).filter(Expense.user_id == current_user.id).all()
    for expense in expenses:
        writer.writerow(["Dépense", expense.date.strftime("%Y-%m-%d"), expense.category, expense.amount, expense.comment or ""])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=finances_export.csv"}
    )


# Route pour export Excel
@app.get("/export/excel")
def export_excel(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Finances"
    
    # En-têtes
    headers = ["Type", "Date", "Catégorie", "Montant", "Commentaire"]
    ws.append(headers)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Revenus
    incomes = db.query(Income).filter(Income.user_id == current_user.id).all()
    for income in incomes:
        ws.append(["Revenu", income.date, income.category, income.amount, ""])
    
    # Dépenses
    expenses = db.query(Expense).filter(Expense.user_id == current_user.id).all()
    for expense in expenses:
        ws.append(["Dépense", expense.date, expense.category, expense.amount, expense.comment or ""])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finances_export.xlsx"}
    )
